import openpyxl # module to read excel files
path = "data1.xlsx"  #path of the excel file
workbook=openpyxl.load_workbook(path)
sheet=workbook.get_sheet_by_name("Sheet4")

# function to read the number of rows in excel
def getRowCount(file,sheetName):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook.get_sheet_by_name(sheetName)
    return(sheet.max_row)

# function to read number of columns in excel
def getColumnCount(file,sheetName):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook.get_sheet_by_name(sheetName)
    return(sheet.max_coloum)

# function to read the excel file
def readData(file,sheetName,rownum,columnno):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook.get_sheet_by_name(sheetName)
    return sheet.cell(row=rownum, column=columnno).value

# function to write data to excel
def writeData(file,sheetName,rownum,columnno,data):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook.get_sheet_by_name(sheetName)
    sheet.cell(row=rownum, column=columnno).value=data
    workbook.save(file)